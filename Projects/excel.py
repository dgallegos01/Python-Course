# here we are going to make a program that will process spreadsheets in excel
# we will work with the 'transactions.xlsx' file

import openpyxl as xl # we are importing the module and giving it an alias 'xl'. you can name any module whatever you want if its easier to reference
wb = xl.load_workbook('transactions.xlsx') # we are calling the file and giving it a variable
sheet = wb['Sheet1'] # we are calling the sheet in the file. this part is case sensitive!
cell = sheet['a1'] # we are accessing each row and column by coordinate. 'a' is the column, 1 is the row
cell = sheet.cell(1,1) # this returns the same cell. its a different way to call the row and column
print(cell.value) # this will print the name of that cell. output: 'transaction_id'
print(sheet.max_row) # this will print the amount of rows in this spreadsheet. should be 4

for row in range(2, sheet.max_row + 1): # '+ 1' lets us include the 4th row. we also need to ignore the 1st row
    print(row)  # making sure we get all 4 rows
    cell = sheet.cell(row, 3) # we are getting the values of column 3
    print(cell.value)
    corrected_price = cell.value * 0.9 # here we are correcting the prices by multiplying it with 0.9
    corrected_price_cell = sheet.cell(row, 4) # accessing column 4
    corrected_price_cell.value = corrected_price # here we are setting the corrected prices to column 4

wb.save('transactions2.xlsx') # we are saving the updated spreadsheet to a new file. the file should show up in the course folder