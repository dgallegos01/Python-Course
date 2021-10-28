# here we are going to add chart to the updated file

import openpyxl as xl 
from openpyxl.chart import BarChart, Reference # we are referencing two classes to make a chart

def process_workbook(filename):
    wb = xl.load_workbook(filename) 
    sheet = wb['Sheet1'] 

    for row in range(2, sheet.max_row + 1): # '+ 1' lets us include the 4th row. we also need to ignore the 1st row
        cell = sheet.cell(row, 3) # we are getting the values of column 3
        corrected_price = cell.value * 0.9 # here we are correcting the prices by multiplying it with 0.9
        corrected_price_cell = sheet.cell(row, 4) # accessing column 4
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4) # we are creating the chart

    chart = BarChart() # we are accessing a chart
    chart.add_data(values) # we are adding data to the chart
    sheet.add_chart(chart, 'e2') # we are adding the chat to column e, row 2

    wb.save(filename)
# go check the file
# there are more things you can do the file via python. go look at the documentation for the package we used